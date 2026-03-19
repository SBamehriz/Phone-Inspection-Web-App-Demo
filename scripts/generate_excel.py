#!/usr/bin/env python3
import sys
import json
import pandas as pd
from datetime import datetime
import os


def apply_header_style(worksheet, color="CC0000"):
    """Apply styled headers with auto column widths."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

    worksheet.row_dimensions[1].height = 28

    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 55)


def grade_fill_color(grade):
    return {
        "A+": "C6EFCE", "A": "C6EFCE",
        "B+": "FFEB9C", "B": "FFEB9C",
        "C": "FFCC99",
    }.get(grade, "EEEEEE")


DEFECT_LABELS = {
    "screen-crack": "Screen Crack",
    "back-damage": "Back Damage",
    "battery-issue": "Battery Issue",
    "camera-malfunction": "Camera Issue",
    "button-stuck": "Button Problems",
    "water-damage": "Water Damage",
    "missing-pen": "Missing Pen",
    "missing-sim-slot": "Missing SIM Slot",
}


# ---------------------------------------------------------------------------
# SINGLE ORDER REPORT
# ---------------------------------------------------------------------------

def build_single_order(writer, order, inspections, timestamp):
    from openpyxl.styles import Font, PatternFill, Alignment

    total = len(inspections)
    completed = sum(1 for i in inspections if i.get("status") == "completed")
    grade_counts = {}
    defect_counts = {}
    for insp in inspections:
        g = insp.get("grade") or "N/A"
        grade_counts[g] = grade_counts.get(g, 0) + 1
        for d in insp.get("defects") or []:
            defect_counts[d] = defect_counts.get(d, 0) + 1

    completion_pct = f"{round(completed / total * 100, 1)}%" if total else "0%"

    # ---- Sheet 1: Order Overview ----------------------------------------
    rows = [
        ["INSPECTION REPORT", ""],
        ["", ""],
        ["ORDER INFORMATION", ""],
        ["Order Number", order.get("orderNumber", "N/A")],
        ["Notes / Description", order.get("notes") or "N/A"],
        ["Expected Quantity", order.get("expectedQuantity", 0)],
        ["Status", (order.get("status") or "").upper()],
        ["Created Date", order.get("createdAt", "N/A")],
        ["Completed Date", order.get("completedAt") or "N/A"],
        ["Report Generated", timestamp],
        ["", ""],
        ["INSPECTION SUMMARY", ""],
        ["Total Devices Scanned", total],
        ["Devices Fully Completed", completed],
        ["Completion Rate", completion_pct],
        ["", ""],
        ["GRADE BREAKDOWN", ""],
    ]
    for g in ["A+", "A", "B+", "B", "C", "N/A"]:
        cnt = grade_counts.get(g, 0)
        pct = f"{round(cnt / total * 100, 1)}%" if total else "0%"
        rows.append([f"Grade {g}", f"{cnt}  ({pct})"])

    rows += [["", ""], ["TOP DEFECTS", ""]]
    if defect_counts:
        for defect, cnt in sorted(defect_counts.items(), key=lambda x: -x[1]):
            pct = f"{round(cnt / total * 100, 1)}%" if total else "0%"
            rows.append([DEFECT_LABELS.get(defect, defect.replace("-", " ").title()), f"{cnt}  ({pct})"])
    else:
        rows.append(["No defects recorded", ""])

    df_overview = pd.DataFrame(rows, columns=["Field", "Value"])
    df_overview.to_excel(writer, sheet_name="Order Overview", index=False)

    ws = writer.sheets["Order Overview"]
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 42
    # Hide generic header row
    ws.row_dimensions[1].height = 1
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", size=1)

    section_titles = {"INSPECTION REPORT", "ORDER INFORMATION", "INSPECTION SUMMARY", "GRADE BREAKDOWN", "TOP DEFECTS"}
    red_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    for row in ws.iter_rows():
        val = row[0].value
        if val in section_titles:
            for cell in row:
                cell.fill = red_fill
                cell.font = Font(bold=True, color="FFFFFF", size=11)
        elif val == "INSPECTION REPORT":
            row[0].font = Font(bold=True, size=14, color="CC0000")

    # ---- Sheet 2: Inspection Details ------------------------------------
    if inspections:
        details = []
        for idx, insp in enumerate(inspections, 1):
            specs = insp.get("phoneSpecs") or {}
            details.append({
                "#": idx,
                "IMEI": insp.get("imei", ""),
                "Brand": specs.get("brand", "Unknown"),
                "Model": specs.get("model", "Unknown"),
                "Storage": specs.get("storage", "Unknown"),
                "Color": specs.get("color", "Unknown"),
                "Grade": insp.get("grade", "Not Graded"),
                "Defects": ", ".join(insp.get("defects") or []).replace("-", " ").title() if insp.get("defects") else "None",
                "Defect Count": len(insp.get("defects") or []),
                "Status": (insp.get("status") or "").upper(),
                "Notes": insp.get("notes") or "",
                "Images": len(insp.get("images") or []),
                "Scanned At": insp.get("scannedAt") or "",
                "Photographed At": insp.get("photographedAt") or "",
                "Completed At": insp.get("completedAt") or "",
            })

        pd.DataFrame(details).to_excel(writer, sheet_name="Inspection Details", index=False)
        ws2 = writer.sheets["Inspection Details"]
        apply_header_style(ws2)

        # Color-code grade cells
        from openpyxl.styles import PatternFill as PF
        grade_col_idx = next((i for i, c in enumerate(ws2[1], 1) if c.value == "Grade"), None)
        if grade_col_idx:
            for row_idx in range(2, len(details) + 2):
                cell = ws2.cell(row=row_idx, column=grade_col_idx)
                cell.fill = PF(start_color=grade_fill_color(cell.value or ""), end_color=grade_fill_color(cell.value or ""), fill_type="solid")

    # ---- Sheet 3: Grade Distribution ------------------------------------
    grade_rows = []
    for g in ["A+", "A", "B+", "B", "C", "N/A"]:
        cnt = grade_counts.get(g, 0)
        pct = round(cnt / total * 100, 1) if total else 0
        grade_rows.append({"Grade": g, "Count": cnt, "Percentage": f"{pct}%"})
    pd.DataFrame(grade_rows).to_excel(writer, sheet_name="Grade Distribution", index=False)
    apply_header_style(writer.sheets["Grade Distribution"])

    # Color grade cells
    from openpyxl.styles import PatternFill as PF
    ws3 = writer.sheets["Grade Distribution"]
    for row_idx in range(2, len(grade_rows) + 2):
        cell = ws3.cell(row=row_idx, column=1)
        color = grade_fill_color(cell.value or "")
        cell.fill = PF(start_color=color, end_color=color, fill_type="solid")

    # ---- Sheet 4: Defect Analysis ---------------------------------------
    if defect_counts:
        defect_rows = [
            {
                "Defect": DEFECT_LABELS.get(d, d.replace("-", " ").title()),
                "Affected Devices": cnt,
                "% of Total Devices": f"{round(cnt / total * 100, 1)}%" if total else "0%",
            }
            for d, cnt in sorted(defect_counts.items(), key=lambda x: -x[1])
        ]
    else:
        defect_rows = [{"Defect": "No defects recorded", "Affected Devices": 0, "% of Total Devices": "0%"}]
    pd.DataFrame(defect_rows).to_excel(writer, sheet_name="Defect Analysis", index=False)
    apply_header_style(writer.sheets["Defect Analysis"])


# ---------------------------------------------------------------------------
# ALL COMPLETED ORDERS REPORT
# ---------------------------------------------------------------------------

def build_all_orders(writer, orders_data, timestamp):
    from openpyxl.styles import PatternFill as PF

    summary_rows = []
    all_device_rows = []
    aggregate_grades = {}
    aggregate_defects = {}

    for item in orders_data:
        order = item["order"]
        inspections = item["inspections"]
        total = len(inspections)
        completed = sum(1 for i in inspections if i.get("status") == "completed")
        gc = {}
        dc = {}
        for insp in inspections:
            g = insp.get("grade") or "N/A"
            gc[g] = gc.get(g, 0) + 1
            aggregate_grades[g] = aggregate_grades.get(g, 0) + 1
            for d in insp.get("defects") or []:
                dc[d] = dc.get(d, 0) + 1
                aggregate_defects[d] = aggregate_defects.get(d, 0) + 1

        top_defect = max(dc, key=dc.get, default=None)
        top_defect_label = DEFECT_LABELS.get(top_defect, top_defect.replace("-", " ").title()) if top_defect else "None"

        summary_rows.append({
            "Order Number": order.get("orderNumber", ""),
            "Notes / Description": order.get("notes") or "",
            "Expected Qty": order.get("expectedQuantity", 0),
            "Devices Scanned": total,
            "Completed": completed,
            "Completion %": f"{round(completed / total * 100, 1)}%" if total else "0%",
            "Grade A+": gc.get("A+", 0),
            "Grade A": gc.get("A", 0),
            "Grade B+": gc.get("B+", 0),
            "Grade B": gc.get("B", 0),
            "Grade C": gc.get("C", 0),
            "Top Defect": top_defect_label,
            "Completed Date": order.get("completedAt") or "N/A",
        })

        for insp in inspections:
            specs = insp.get("phoneSpecs") or {}
            all_device_rows.append({
                "Order Number": order.get("orderNumber", ""),
                "IMEI": insp.get("imei", ""),
                "Brand": specs.get("brand", "Unknown"),
                "Model": specs.get("model", "Unknown"),
                "Storage": specs.get("storage", "Unknown"),
                "Color": specs.get("color", "Unknown"),
                "Grade": insp.get("grade", "Not Graded"),
                "Defects": ", ".join(insp.get("defects") or []).replace("-", " ").title() if insp.get("defects") else "None",
                "Status": (insp.get("status") or "").upper(),
                "Notes": insp.get("notes") or "",
                "Completed At": insp.get("completedAt") or "",
            })

    # Sheet 1: Orders Summary
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Orders Summary", index=False)
    apply_header_style(writer.sheets["Orders Summary"])

    # Sheet 2: All Devices
    if all_device_rows:
        pd.DataFrame(all_device_rows).to_excel(writer, sheet_name="All Devices", index=False)
        ws = writer.sheets["All Devices"]
        apply_header_style(ws)
        grade_col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Grade"), None)
        if grade_col:
            for row_idx in range(2, len(all_device_rows) + 2):
                cell = ws.cell(row=row_idx, column=grade_col)
                color = grade_fill_color(cell.value or "")
                cell.fill = PF(start_color=color, end_color=color, fill_type="solid")

    # Sheet 3: Grade Summary
    total_devices = len(all_device_rows)
    grade_rows = []
    for g in ["A+", "A", "B+", "B", "C", "N/A"]:
        cnt = aggregate_grades.get(g, 0)
        pct = round(cnt / total_devices * 100, 1) if total_devices else 0
        grade_rows.append({"Grade": g, "Total Count": cnt, "Percentage": f"{pct}%"})
    pd.DataFrame(grade_rows).to_excel(writer, sheet_name="Grade Summary", index=False)
    ws3 = writer.sheets["Grade Summary"]
    apply_header_style(ws3)
    for row_idx in range(2, len(grade_rows) + 2):
        cell = ws3.cell(row=row_idx, column=1)
        color = grade_fill_color(cell.value or "")
        cell.fill = PF(start_color=color, end_color=color, fill_type="solid")

    # Sheet 4: Defect Summary
    if aggregate_defects:
        defect_rows = [
            {
                "Defect": DEFECT_LABELS.get(d, d.replace("-", " ").title()),
                "Total Occurrences": cnt,
                "% of All Devices": f"{round(cnt / total_devices * 100, 1)}%" if total_devices else "0%",
            }
            for d, cnt in sorted(aggregate_defects.items(), key=lambda x: -x[1])
        ]
    else:
        defect_rows = [{"Defect": "No defects recorded", "Total Occurrences": 0, "% of All Devices": "0%"}]
    pd.DataFrame(defect_rows).to_excel(writer, sheet_name="Defect Summary", index=False)
    apply_header_style(writer.sheets["Defect Summary"])


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

def generate_excel_report(data_json_path, output_path):
    with open(data_json_path, "r") as f:
        data = json.load(f)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    timestamp = data.get("timestamp", datetime.now().isoformat())

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if data.get("mode") == "all":
            build_all_orders(writer, data["orders"], timestamp)
        else:
            build_single_order(writer, data["order"], data["inspections"], timestamp)

    print(f"Excel report generated successfully: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_excel.py <data_json_path> <output_path>")
        sys.exit(1)

    try:
        generate_excel_report(sys.argv[1], sys.argv[2])
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        sys.exit(1)
